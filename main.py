from datetime import datetime, timedelta

### These libraries needs to be uploaded ###
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook
import pandas as pd
import sqlite3

database_dosya_yolu = "data.csv"
degiskenlercsv_ortak_dosya_yolu = "variables.csv"
utilizasyonraporu_cikti_dosyasi_yolu = "Utilizasyon.xlsx"
sql_database = 'monthlyrobots.db'



conn = sqlite3.connect(sql_database)
cursor = conn.cursor()
cursor.execute('''
CREATE TABLE IF NOT EXISTS monthlyrobots (
    robot_name TEXT,
    date DATETIME,
    percentage TEXT
)
''')
conn.commit()
conn.close()




with open(degiskenlercsv_ortak_dosya_yolu, "r", encoding="utf-8") as dosya:
    degisken_satirlari = dosya.readlines()

with open(database_dosya_yolu, "r", encoding="utf-8") as dosya:
    veri_satirlari = dosya.readlines()

day_interval = int(degisken_satirlari[1].split(";")[1])
min_free_time_minutes = int(degisken_satirlari[2].split(";")[1])


i = 0
while i < len(veri_satirlari):
    if veri_satirlari[-(i+1)].split(",")[11].strip() != "":
        lastdate = datetime.strptime(veri_satirlari[-(i+1)].split(",")[11].strip(), "%m/%d/%Y %I:%M:%S %p")
        startdate = lastdate - timedelta(days=day_interval)

        break
    i += 1

titles = veri_satirlari[0].split(",")
duzenlenmis_veri_satirlari = []

# ====== Datayi duzenliyoruz ve sadelestiriyoruz. ======

# Process;Machine;Hostname;Host Identity;Job type;Runtime type/license;State;Priority;Started (absolute);Ended (absolute);Source;Created (absolute)
duzenlenmis_veri_satirlari.append(";".join([titles[1], 
                                            titles[3], 
                                            titles[4], 
                                            titles[6], 
                                            titles[7], 
                                            titles[8], 
                                            titles[9], 
                                            titles[10], 
                                            titles[11], 
                                            titles[12], 
                                            titles[13], 
                                            titles[14][:-1]]))


counter = 0
for satir in veri_satirlari:
    if counter > 0:
        satir_bolunmus = satir.strip().split(",")
        yeni_satir = []
        for eleman in satir_bolunmus:
            yeni_satir.append("")
            yeni_satir.append(eleman)

        try:
            tempdate = datetime.strptime(yeni_satir[23], "%m/%d/%Y %I:%M:%S %p")
            if tempdate > startdate:
                duzenlenmis_veri_satirlari.append(";".join([yeni_satir[3], 
                                                            yeni_satir[7], 
                                                            yeni_satir[9], 
                                                            yeni_satir[13], 
                                                            yeni_satir[15], 
                                                            yeni_satir[17], 
                                                            yeni_satir[19], 
                                                            yeni_satir[21], 
                                                            yeni_satir[23], 
                                                            yeni_satir[25], 
                                                            yeni_satir[27], 
                                                            yeni_satir[29]]))

        except:
            print(yeni_satir[23])
            

    counter += 1


# with open("sorgulanmisveriler.csv", "w", encoding="utf-8-sig") as yeni_dosya:
#     for satir in duzenlenmis_veri_satirlari:
#         yeni_dosya.write(satir + "\n")
print(str(startdate) + " tarihinden " + str(lastdate) + " tarihine kadar olan veriler excel e aktarildi.")


# ============ Verileri robotlara atİyoruz. ============

# ilk olarak bir json veritabanİmİz olacak bu veritabanİnda ROBOT:{name:SeherGida;Status:Success} seklinde veriler tutulacak.
# ilk 1. boyutta robot isimlerinin bulunacagi listeleri olusturmamiz gerekiyor.

robots = []
utilization = {}
counter = 0


# ilk boyutta robot isimlerini topluyoruz
for i in duzenlenmis_veri_satirlari:
    if counter > 0:
        robotname = i.split(";")[3]
        if robotname[11:] not in robots and robotname[11:] != "":
            robots.append(robotname[11:])
    counter += 1

# Robotlar listede gordukleri siraya gore ekleniyordu. Asagidaki fonksiyona gore Robotlarin numaralarina gore atanmasi saglandi. 
def sort_robots(robot_list):
    def robot_key(robot):
        if robot == "ROBOT":
            return 1
        else:
            return int(robot.replace("ROBOT", "")) * 10

    # Robotlari numaralarina gore siraliyoruz
    sorted_list = sorted(robot_list, key=robot_key)
    return sorted_list

robots = sort_robots(robots)

for robot_name in robots:
    utilization[robot_name] = {"processes": [], 
                               "efficiency": [],
                               "available":[]}

# ikinci boyutta robot surelerini dolduruyoruz
counter = 0
for i in duzenlenmis_veri_satirlari:
    if counter > 0:
        data = i.split(";")
        name = data[0]
        robotname = data[3]
        starteddate = data[8]
        endeddate = data[9]
        processresult = data[6]
        processstarter = data[10]

        if robotname[11:] in utilization:
            process_data = {
                "name": name,
                "starteddate": starteddate,
                "endeddate": endeddate,
                "processresult": processresult,
                "processstarter": processstarter
            }
            utilization[robotname[11:]]["processes"].append(process_data)

    counter += 1


def format_time(minutes):
    """This Function Gives You Day, Hour, Minutes"""
    days = int(minutes // (24 * 60))
    hours = int((minutes % (24 * 60)) // 60)
    mins = int(minutes % 60)
    
    if days > 0:
        return f"{days} day, {hours} hours, {mins} mins"
    elif hours > 0:
        return f"{hours} hours, {mins} mins"
    else:
        return f"{mins} mins"


def calculate_time_difference_in_minutes(start_date_str, end_date_str):
    start_date = datetime.strptime(start_date_str, "%m/%d/%Y %I:%M:%S %p")
    end_date = datetime.strptime(end_date_str, "%m/%d/%Y %I:%M:%S %p")

    time_difference = end_date - start_date
    time_difference_in_minutes = time_difference.total_seconds() / 60

    return time_difference_in_minutes

def calculate_daily_free_times(free_intervals):
    """Calculate daily free times from free intervals"""
    daily_free_times = {}

    for interval in free_intervals:
        start = datetime.strptime(interval['start'], "%m/%d/%Y %I:%M:%S %p")
        end = datetime.strptime(interval['end'], "%m/%d/%Y %I:%M:%S %p")
        current = start

        while current < end:
            day_end = datetime(current.year, current.month, current.day, 23, 59, 59)
            if day_end > end:
                day_end = end

            free_time = (day_end - current).total_seconds() / 60
            day_key = current.strftime("%d/%m/%Y")

            if day_key in daily_free_times:
                daily_free_times[day_key] += free_time
            else:
                daily_free_times[day_key] = free_time

            current = day_end + timedelta(seconds=1)  # Move to the next day

    return daily_free_times

def save_data_to_db(robot, tarih, yuzde):
    conn = sqlite3.connect('monthlyrobots.db')
    cursor = conn.cursor()

    # Tabloyu oluşturma (eğer yoksa)
    cursor.execute('''CREATE TABLE IF NOT EXISTS monthlyrobots
                      (robot_name TEXT, date DATETIME, percentage TEXT)''')

    # Aynı robot ve tarih kombinasyonuna sahip bir kayıt olup olmadığını kontrol etme
    cursor.execute("SELECT * FROM monthlyrobots WHERE robot_name = ? AND date = ?", (robot, tarih))
    data = cursor.fetchone()

    if data is None:
        # Eger kayıt yoksa yeni veriyi ekle
        cursor.execute("INSERT INTO monthlyrobots (robot_name, date, percentage) VALUES (?, ?, ?)", (robot, tarih, yuzde))
        conn.commit()
    else:
        print(f"{robot} için {tarih} tarihinde zaten bir kayıt mevcut.")

    conn.close()

    
bold_font = Font(bold=True)
availabletimes = ""

wb = Workbook()

for robotprocesses in utilization:
    success_counter = 0
    faulted_counter = 0
    running_counter = 0
    pending_counter = 0
    stopped_counter = 0
    calc_successful = 0
    robot_free_intervals = []
    
    ws = wb.create_sheet(robotprocesses)
    ws.append(["Process", "Started (absolute)", "Ended (absolute)", "Time", "Status", "Trigger"])
    for cell in ws[1]:
        cell.font = bold_font

    sorted_processes = sorted(utilization[robotprocesses]["processes"], key=lambda x: datetime.strptime(x["starteddate"], "%m/%d/%Y %I:%M:%S %p"))
    for idx, process in enumerate(sorted_processes):
        temp = ""
        if process["processresult"] == "Successful":
            success_counter += 1
            temp = calculate_time_difference_in_minutes(process["starteddate"], process["endeddate"])
            calc_successful += temp
        elif process["processresult"] == "Faulted":
            faulted_counter += 1
        elif process["processresult"] == "Running":
            running_counter += 1
        elif process["processresult"] == "Pending":
            pending_counter += 1
        elif process["processresult"] == "Stopped":
            stopped_counter += 1
        if idx > 0:
            previous_process = sorted_processes[idx - 1]
            end_prev = datetime.strptime(previous_process["endeddate"], "%m/%d/%Y %I:%M:%S %p")
            start_current = datetime.strptime(process["starteddate"], "%m/%d/%Y %I:%M:%S %p")
            free_time = calculate_time_difference_in_minutes(end_prev.strftime("%m/%d/%Y %I:%M:%S %p"), start_current.strftime("%m/%d/%Y %I:%M:%S %p"))
            if end_prev < start_current and free_time >= min_free_time_minutes:
                robot_free_intervals.append({
                    "start": end_prev.strftime("%m/%d/%Y %I:%M:%S %p"),
                    "end": start_current.strftime("%m/%d/%Y %I:%M:%S %p")
                })
        tempcalc = calc_successful / (day_interval * 24 * 60) * 100
        tempcalc = round(tempcalc, 2)
        if type(temp) != str:
            temp = str(round(temp, 2)) + " minute"
        else:
            temp = ""
        ws.append([process["name"], process["starteddate"], process["endeddate"], temp, process["processresult"], process["processstarter"]])

    utilization[robotprocesses]["efficiency"] = {
        "Total Consumed Time": str(calc_successful),
        "Successful Counter": success_counter, 
        "Faulted Counter": faulted_counter, 
        "Running Counter": running_counter,
        "Pending Counter": pending_counter,
        "Efficiency": "%" + str(tempcalc)
    }
    utilization[robotprocesses]["available"] = robot_free_intervals
    ws.append([])
    ws.append([])
    save_data_to_db(robotprocesses, (datetime.strptime(process["starteddate"], "%m/%d/%Y %I:%M:%S %p")).strftime("%d/%m/%Y"), "%" + str(tempcalc))

    ws.append(["Total Consumed Time", str(round(calc_successful, 3)) + " minute"])
    ws[ws.max_row][0].font = bold_font
    ws.append(["Successful Counter", success_counter])
    ws[ws.max_row][0].font = bold_font
    ws.append(["Faulted Counter", faulted_counter])
    ws[ws.max_row][0].font = bold_font
    ws.append(["Running Counter", running_counter])
    ws[ws.max_row][0].font = bold_font    
    ws.append(["Pending Counter", pending_counter])
    ws[ws.max_row][0].font = bold_font
    ws.append(["Efficiency", "%" + str(tempcalc)])
    ws[ws.max_row][0].font = bold_font
    
    ws.append([])
    ws.append([])

    if utilization[robotprocesses]["available"]:
        ws.append(["Available Start Times", "Available End Times", "Free Times"])
        ws[ws.max_row][0].font = bold_font
        ws[ws.max_row][1].font = bold_font
        ws[ws.max_row][2].font = bold_font

    for gapdata in utilization[robotprocesses]["available"]:
        gap = format_time(calculate_time_difference_in_minutes(gapdata["start"], gapdata["end"]))
        ws.append([gapdata["start"], gapdata["end"], gap])
    
    daily_free_times = calculate_daily_free_times(utilization[robotprocesses]["available"])
    ws.append([])
    ws.append([])
    ws.append(["Daily Free Times", "Total Free Times"])
    ws[ws.max_row][0].font = bold_font
    ws[ws.max_row][1].font = bold_font
    for day, free_time in daily_free_times.items():
        ws.append([day, format_time(free_time)])



ws = wb.create_sheet("Efficiency by Dates")
connection = sqlite3.connect(sql_database)
df = pd.read_sql_query("SELECT * FROM monthlyrobots", connection)
connection.close()

df.columns = df.columns.str.lower().str.strip()
df['date'] = pd.to_datetime(df['date'], format='%d/%m/%Y')
df['percentage'] = df['percentage'].str.replace('%', '').astype(float)
df['date'] = df['date'].dt.date
pivot_df = df.pivot(index='date', columns='robot_name', values='percentage')

if not all(date == df['date'][0] for date in df['date']):
    for r in dataframe_to_rows(pivot_df, index=True, header=True):
        ws.append(r)

    ws.delete_rows(ws.max_row-4, 1)
    # Graph creator
    chart = LineChart()
    chart.title = "Robot Efficiencies"
    chart.style = 10
    chart.y_axis.title = 'Efficiency (%)'
    chart.x_axis.title = 'Date'

    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)

    dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.set_categories(dates)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "A"+str(ws.max_row+2))
else:
    wb.remove(wb["Efficiency by Dates"])

wb.remove(wb["Sheet"])
wb.save(utilizasyonraporu_cikti_dosyasi_yolu)


